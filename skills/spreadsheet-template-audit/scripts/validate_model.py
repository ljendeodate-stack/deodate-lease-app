#!/usr/bin/env python3
"""
Validate a DEODATE Lease Obligation Analysis workbook against 25+ business
logic assertions.

Runs after both population and repair workflows. Checks date math, dollar
tie-outs, formula integrity, NRC dates, escalation plausibility, and
cross-sheet linkage.

Usage:
    python validate_model.py <workbook_path> [--sheet "Lease Schedule (2)"]

Returns JSON with pass/fail status for each assertion.
Can also be imported: from validate_model import validate
"""

import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
import calendar

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl required. Install: pip install openpyxl")
    sys.exit(1)


# ── Sheet configurations ──────────────────────────────────────────────────────

CONFIGS = {
    "Lease Schedule": {
        "first_row": 38, "last_row": 161, "total_row": 162,
        "sf_cell": "C7", "commence_cell": "C8", "term_cell": "C10",
        "expire_cell": "C11", "eff_date_cell": "C12", "eff_month_cell": "C13",
        "base_rent_cell": "C21",
        "nnn_cells": ["C22", "C23", "C24", "C25", "C26"],
        "esc_cells": ["F5", "F6", "F7", "F8", "F9", "F10"],
        "abat_dur_cell": "C29", "abat_pct_cell": "C32", "abat_amt_cell": "C33",
        "abat_start_cell": "C30", "abat_end_cell": "C31",
        "discount_rate_cell": "J2",
        "nrc_amount_range": ("F", 14, 26), "nrc_date_range": ("G", 14, 26),
        "helper_cols": ["T", "U", "V", "W"],
        "total_col_N": "N",
        "annual_summary_sheet": "Annual Summary",
    },
    "Lease Schedule (2)": {
        "first_row": 75, "last_row": 198, "total_row": 199,
        "sf_cell": "C7", "commence_cell": "C8", "term_cell": "C10",
        "expire_cell": "C11", "eff_date_cell": "C12", "eff_month_cell": "C13",
        "base_rent_cell": "C21",
        "nnn_cells": ["C22", "C23", "C24", "C25", "C26"],
        "esc_cells": ["C29", "C30", "C31", "C32", "C33", "C34"],
        "abat_dur_cell": "C37", "abat_pct_cell": "C40", "abat_amt_cell": "C41",
        "abat_start_cell": "C38", "abat_end_cell": "C39",
        "discount_rate_cell": "F7",
        "nrc_amount_range": ("C", 53, 65), "nrc_date_range": ("D", 53, 65),
        "helper_cols": ["T", "U", "V", "W"],
        "total_col_N": "N",
        "annual_summary_sheet": "Annual Summary",
    },
}


def cell_val(ws, ref):
    """Get cell value, returning None for empty cells."""
    v = ws[ref].value
    return v


def cell_num(ws, ref, default=0):
    """Get numeric cell value, coercing None to default."""
    v = ws[ref].value
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def cell_date(ws, ref):
    """Get date cell value."""
    v = ws[ref].value
    if isinstance(v, datetime):
        return v
    return None


def add_months(dt, months):
    """Add N months to a date."""
    month = dt.month - 1 + months
    year = dt.year + month // 12
    month = month % 12 + 1
    return datetime(year, month, 1)


def end_of_month(dt):
    """Last day of month."""
    last_day = calendar.monthrange(dt.year, dt.month)[1]
    return datetime(dt.year, dt.month, last_day)


class Assertion:
    def __init__(self, id, name, severity, passed, detail):
        self.id = id
        self.name = name
        self.severity = severity  # CRITICAL, HIGH, MEDIUM, LOW
        self.passed = passed
        self.detail = detail

    def to_dict(self):
        return {
            "id": self.id,
            "name": self.name,
            "severity": self.severity,
            "passed": self.passed,
            "detail": self.detail,
        }


def validate(workbook_path, sheet_name=None):
    """Run all assertions against the workbook. Returns structured results."""
    wb_path = Path(workbook_path)
    if not wb_path.exists():
        return {"status": "error", "message": f"File not found: {wb_path}"}

    # Load both data_only (for calculated values) and formula mode
    wb_data = load_workbook(str(wb_path), data_only=True)
    wb_formula = load_workbook(str(wb_path), data_only=False)

    # Detect sheet
    if sheet_name is None:
        for name in CONFIGS:
            if name in wb_data.sheetnames:
                ws = wb_data[name]
                if cell_num(ws, "C21") > 0:
                    sheet_name = name
                    break
        if sheet_name is None:
            sheet_name = next((n for n in CONFIGS if n in wb_data.sheetnames), None)
    if sheet_name is None or sheet_name not in CONFIGS:
        return {"status": "error", "message": f"No valid sheet found or configured"}

    cfg = CONFIGS[sheet_name]
    ws = wb_data[sheet_name]
    ws_f = wb_formula[sheet_name]
    assertions = []
    a_id = 0

    # ── 1. Rentable SF > 0 ───────────────────────────────────────────────
    a_id += 1
    sf = cell_num(ws, cfg["sf_cell"])
    assertions.append(Assertion(a_id, "Rentable SF is positive", "CRITICAL",
        sf > 0, f"SF = {sf}"))

    # ── 2. Lease term is plausible (1–600) ───────────────────────────────
    a_id += 1
    term = cell_num(ws, cfg["term_cell"])
    assertions.append(Assertion(a_id, "Lease term is plausible (1–600 months)", "HIGH",
        1 <= term <= 600, f"Term = {term}"))

    # ── 3. Commencement + term ≈ expiration ──────────────────────────────
    a_id += 1
    commence = cell_date(ws, cfg["commence_cell"])
    expire = cell_date(ws, cfg["expire_cell"])
    if commence and expire and term > 0:
        expected_exp = end_of_month(add_months(commence, int(term) - 1))
        diff_days = abs((expire - expected_exp).days)
        assertions.append(Assertion(a_id, "Commencement + term ≈ expiration date", "HIGH",
            diff_days <= 31, f"Expected ≈ {expected_exp.strftime('%Y-%m-%d')}, actual = {expire.strftime('%Y-%m-%d')}, diff = {diff_days} days"))
    else:
        assertions.append(Assertion(a_id, "Commencement + term ≈ expiration date", "HIGH",
            False, "Missing commencement, term, or expiration date"))

    # ── 4. Base rent is positive ─────────────────────────────────────────
    a_id += 1
    base_rent = cell_num(ws, cfg["base_rent_cell"])
    assertions.append(Assertion(a_id, "Base rent Year 1 is positive", "CRITICAL",
        base_rent > 0, f"Base rent = ${base_rent:,.2f}"))

    # ── 5. All escalation rates 0–20% ───────────────────────────────────
    a_id += 1
    esc_ok = True
    esc_detail = []
    for cell in cfg["esc_cells"]:
        rate = cell_num(ws, cell)
        if rate < 0 or rate > 0.20:
            esc_ok = False
            esc_detail.append(f"{cell}={rate:.1%}")
    assertions.append(Assertion(a_id, "All escalation rates within 0–20%", "MEDIUM",
        esc_ok, f"Out of range: {', '.join(esc_detail)}" if esc_detail else "All within range"))

    # ── 6. No zero escalation rates (unless intentional) ────────────────
    a_id += 1
    zero_esc = [cell for cell in cfg["esc_cells"] if cell_num(ws, cell) == 0]
    assertions.append(Assertion(a_id, "No zero escalation rates (warning only)", "LOW",
        len(zero_esc) == 0, f"Zero rates at: {', '.join(zero_esc)}" if zero_esc else "All non-zero"))

    # ── 7. NNN rates are non-negative ───────────────────────────────────
    a_id += 1
    neg_nnn = [cell for cell in cfg["nnn_cells"] if cell_num(ws, cell) < 0]
    assertions.append(Assertion(a_id, "All NNN Year 1 rates are non-negative", "HIGH",
        len(neg_nnn) == 0, f"Negative rates at: {', '.join(neg_nnn)}" if neg_nnn else "All non-negative"))

    # ── 8. Abatement duration < term ────────────────────────────────────
    a_id += 1
    abat_dur = cell_num(ws, cfg["abat_dur_cell"])
    assertions.append(Assertion(a_id, "Abatement duration < lease term", "HIGH",
        abat_dur < term, f"Abatement = {abat_dur} months, term = {term} months"))

    # ── 9. Abatement amount ≈ base rent × abatement % ──────────────────
    a_id += 1
    abat_pct = cell_num(ws, cfg["abat_pct_cell"])
    abat_amt = cell_num(ws, cfg["abat_amt_cell"])
    if abat_dur > 0:
        expected_amt = base_rent * abat_pct
        diff_pct = abs(abat_amt - expected_amt) / max(expected_amt, 1) if expected_amt > 0 else 0
        assertions.append(Assertion(a_id, "Abatement amount ≈ base rent × abatement %", "MEDIUM",
            diff_pct < 0.01, f"Expected ${expected_amt:,.2f}, actual ${abat_amt:,.2f}, diff = {diff_pct:.1%}"))
    else:
        assertions.append(Assertion(a_id, "Abatement amount ≈ base rent × abatement %", "MEDIUM",
            True, "No abatement period — skipped"))

    # ── 10. Abatement dates consistent with duration ────────────────────
    a_id += 1
    if abat_dur > 0 and commence:
        abat_start = cell_date(ws, cfg["abat_start_cell"])
        abat_end = cell_date(ws, cfg["abat_end_cell"])
        if abat_start and abat_end:
            expected_end = end_of_month(add_months(abat_start, int(abat_dur) - 1))
            diff_days = abs((abat_end - expected_end).days)
            assertions.append(Assertion(a_id, "Abatement end date consistent with start + duration", "MEDIUM",
                diff_days <= 1, f"Expected {expected_end.strftime('%Y-%m-%d')}, actual {abat_end.strftime('%Y-%m-%d')}"))
        else:
            assertions.append(Assertion(a_id, "Abatement end date consistent with start + duration", "MEDIUM",
                False, "Missing abatement start or end date"))
    else:
        assertions.append(Assertion(a_id, "Abatement end date consistent with start + duration", "MEDIUM",
            True, "No abatement — skipped"))

    # ── 11. Discount rate is plausible (0–25%) ──────────────────────────
    a_id += 1
    disc_rate = cell_num(ws, cfg["discount_rate_cell"])
    assertions.append(Assertion(a_id, "Discount rate is plausible (0–25%)", "MEDIUM",
        0 < disc_rate <= 0.25, f"Discount rate = {disc_rate:.1%}"))

    # ── 12. Effective month # is between 1 and term ─────────────────────
    a_id += 1
    eff_month = cell_num(ws, cfg["eff_month_cell"])
    assertions.append(Assertion(a_id, "Effective month # is between 1 and term", "HIGH",
        1 <= eff_month <= term, f"Effective month = {eff_month}, term = {term}"))

    # ── 13. Schedule row count matches term ─────────────────────────────
    a_id += 1
    first_row = cfg["first_row"]
    last_row = cfg["last_row"]
    sched_months = 0
    for r in range(first_row, last_row + 1):
        if ws.cell(row=r, column=1).value is not None:
            sched_months += 1
        else:
            break
    assertions.append(Assertion(a_id, "Schedule row count matches term", "HIGH",
        sched_months == int(term), f"Schedule has {sched_months} rows, term = {int(term)}"))

    # ── 14. TOTAL row col N = SUM of schedule col N ─────────────────────
    a_id += 1
    total_row = cfg["total_row"]
    total_n = cell_num(ws, f"N{total_row}")
    sum_n = sum(cell_num(ws, f"N{r}") for r in range(first_row, last_row + 1))
    if total_n > 0:
        diff_pct = abs(total_n - sum_n) / total_n
        assertions.append(Assertion(a_id, "TOTAL row N ties to SUM of schedule N", "CRITICAL",
            diff_pct < 0.001, f"TOTAL = ${total_n:,.2f}, SUM = ${sum_n:,.2f}, diff = {diff_pct:.4%}"))
    else:
        assertions.append(Assertion(a_id, "TOTAL row N ties to SUM of schedule N", "CRITICAL",
            sum_n == 0, f"TOTAL = ${total_n:,.2f}, SUM = ${sum_n:,.2f}"))

    # ── 15. TOTAL row col F = SUM of schedule col F ─────────────────────
    a_id += 1
    total_f = cell_num(ws, f"F{total_row}")
    sum_f = sum(cell_num(ws, f"F{r}") for r in range(first_row, last_row + 1))
    if total_f > 0:
        diff_pct = abs(total_f - sum_f) / total_f
        assertions.append(Assertion(a_id, "TOTAL row F ties to SUM of schedule F", "HIGH",
            diff_pct < 0.001, f"TOTAL = ${total_f:,.2f}, SUM = ${sum_f:,.2f}"))
    else:
        assertions.append(Assertion(a_id, "TOTAL row F ties to SUM of schedule F", "HIGH",
            True, "Both zero"))

    # ── 16. P{first_row} ≈ SUM(N all) + S{first_row} ───────────────────
    a_id += 1
    p_first = cell_num(ws, f"P{first_row}")
    s_first = cell_num(ws, f"S{first_row}")
    expected_p = sum_n + s_first
    if expected_p != 0:
        diff_pct = abs(p_first - expected_p) / abs(expected_p)
        assertions.append(Assertion(a_id, "Obligation Remaining (P1) = SUM(N) + OTC(S1)", "HIGH",
            diff_pct < 0.001, f"P{first_row} = ${p_first:,.2f}, expected = ${expected_p:,.2f}"))
    else:
        assertions.append(Assertion(a_id, "Obligation Remaining (P1) = SUM(N) + OTC(S1)", "HIGH",
            True, "Both zero"))

    # ── 17. T-W helper columns populated ────────────────────────────────
    a_id += 1
    tw_populated = True
    tw_detail = []
    for col in cfg["helper_cols"]:
        cell_ref = f"{col}{first_row}"
        v = ws_f[cell_ref].value
        if v is None or (isinstance(v, (int, float)) and v == 0):
            tw_populated = False
            tw_detail.append(f"{cell_ref} is empty")
    assertions.append(Assertion(a_id, "Helper columns T–W are populated with formulas", "CRITICAL",
        tw_populated, "; ".join(tw_detail) if tw_detail else "All populated"))

    # ── 18. T-W formulas reference correct abatement cell ───────────────
    a_id += 1
    if tw_populated:
        t_formula = str(ws_f[f"T{first_row}"].value or "")
        abat_ref = cfg["abat_dur_cell"].replace("$", "")
        assertions.append(Assertion(a_id, "T-W formulas reference correct abatement cell", "HIGH",
            abat_ref.replace("$", "") in t_formula.replace("$", ""),
            f"T{first_row} formula: {t_formula[:80]}..."))
    else:
        assertions.append(Assertion(a_id, "T-W formulas reference correct abatement cell", "HIGH",
            False, "T-W columns are empty — cannot verify formula references"))

    # ── 19. NRC dates are date serials (not text) ───────────────────────
    a_id += 1
    nrc_col, nrc_start, nrc_end = cfg["nrc_date_range"]
    amt_col, _, _ = cfg["nrc_amount_range"]
    text_dates = []
    blank_dates = []
    for r in range(nrc_start, nrc_end + 1):
        date_val = ws.cell(row=r, column=ws[f"{nrc_col}1"].column).value
        amt_val = ws.cell(row=r, column=ws[f"{amt_col}1"].column).value
        if amt_val is not None and amt_val != 0:
            if date_val is None or date_val == "":
                blank_dates.append(f"Row {r}")
            elif isinstance(date_val, str):
                text_dates.append(f"Row {r}: '{date_val}'")
    nrc_ok = len(text_dates) == 0 and len(blank_dates) == 0
    detail_parts = []
    if text_dates:
        detail_parts.append(f"Text dates: {'; '.join(text_dates)}")
    if blank_dates:
        detail_parts.append(f"Blank dates: {'; '.join(blank_dates)}")
    assertions.append(Assertion(a_id, "All NRC dates are date serials (not text or blank)", "LOW",
        nrc_ok, "; ".join(detail_parts) if detail_parts else "All valid date serials"))

    # ── 20. Abatement column G sums to negative of expected total ───────
    a_id += 1
    if abat_dur > 0:
        sum_g = sum(cell_num(ws, f"G{r}") for r in range(first_row, last_row + 1))
        expected_abat_total = -(base_rent * abat_pct * abat_dur)
        if expected_abat_total != 0:
            diff_pct = abs(sum_g - expected_abat_total) / abs(expected_abat_total)
            assertions.append(Assertion(a_id, "Abatement column G total ties to expected", "MEDIUM",
                diff_pct < 0.05, f"SUM(G) = ${sum_g:,.2f}, expected ≈ ${expected_abat_total:,.2f}"))
        else:
            assertions.append(Assertion(a_id, "Abatement column G total ties to expected", "MEDIUM",
                True, "Expected zero, got zero"))
    else:
        assertions.append(Assertion(a_id, "Abatement column G total ties to expected", "MEDIUM",
            True, "No abatement — skipped"))

    # ── 21. No formula errors in schedule range ─────────────────────────
    a_id += 1
    error_cells = []
    error_values = {"#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!"}
    for r in range(first_row, last_row + 1):
        for c in range(1, 20):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() in error_values:
                from openpyxl.utils import get_column_letter
                error_cells.append(f"{get_column_letter(c)}{r}={v}")
    assertions.append(Assertion(a_id, "No formula errors in schedule range", "CRITICAL",
        len(error_cells) == 0, f"{len(error_cells)} errors: {', '.join(error_cells[:5])}" if error_cells else "Clean"))

    # ── 22. Month # column is sequential 1–N ────────────────────────────
    a_id += 1
    month_seq_ok = True
    month_detail = ""
    for i, r in enumerate(range(first_row, first_row + sched_months)):
        m = cell_num(ws, f"C{r}")
        if int(m) != i + 1:
            month_seq_ok = False
            month_detail = f"C{r} = {int(m)}, expected {i + 1}"
            break
    assertions.append(Assertion(a_id, "Month # column is sequential 1–N", "HIGH",
        month_seq_ok, month_detail if month_detail else "Sequential"))

    # ── 23. Year # column increments every 12 months ────────────────────
    a_id += 1
    year_ok = True
    year_detail = ""
    for i, r in enumerate(range(first_row, first_row + sched_months)):
        y = cell_num(ws, f"D{r}")
        expected_y = (i // 12) + 1
        if int(y) != expected_y:
            year_ok = False
            year_detail = f"D{r} = {int(y)}, expected {expected_y}"
            break
    assertions.append(Assertion(a_id, "Year # increments every 12 months", "HIGH",
        year_ok, year_detail if year_detail else "Correct"))

    # ── 24. Annual Summary grand total ties to schedule TOTAL ───────────
    a_id += 1
    as_name = cfg["annual_summary_sheet"]
    if as_name in wb_data.sheetnames:
        as_ws = wb_data[as_name]
        as_total = cell_num(as_ws, "G13")
        if total_n > 0 and as_total > 0:
            diff_pct = abs(as_total - total_n) / total_n
            assertions.append(Assertion(a_id, "Annual Summary grand total ties to schedule TOTAL (N)", "HIGH",
                diff_pct < 0.001, f"Annual Summary G13 = ${as_total:,.2f}, Schedule TOTAL N = ${total_n:,.2f}"))
        elif total_n > 0 and as_total == 0:
            assertions.append(Assertion(a_id, "Annual Summary grand total ties to schedule TOTAL (N)", "HIGH",
                False, f"Annual Summary = $0, Schedule TOTAL = ${total_n:,.2f} — likely wrong sheet reference"))
        else:
            assertions.append(Assertion(a_id, "Annual Summary grand total ties to schedule TOTAL (N)", "HIGH",
                True, "Both zero or schedule is blank"))
    else:
        assertions.append(Assertion(a_id, "Annual Summary grand total ties to schedule TOTAL (N)", "HIGH",
            False, f"Sheet '{as_name}' not found"))

    # ── 25. N = F + M for all schedule rows ─────────────────────────────
    a_id += 1
    n_eq_fm_ok = True
    n_detail = ""
    for r in range(first_row, first_row + sched_months):
        f_val = cell_num(ws, f"F{r}")
        m_val = cell_num(ws, f"M{r}")
        n_val = cell_num(ws, f"N{r}")
        if abs(n_val - (f_val + m_val)) > 0.01:
            n_eq_fm_ok = False
            n_detail = f"Row {r}: N={n_val:.2f}, F+M={f_val + m_val:.2f}"
            break
    assertions.append(Assertion(a_id, "N = F + M (Total Obligation = Base Rent + NNN) for all rows", "CRITICAL",
        n_eq_fm_ok, n_detail if n_detail else "All rows tie"))

    # ── Compile results ──────────────────────────────────────────────────
    results = [a.to_dict() for a in assertions]
    passed = sum(1 for a in assertions if a.passed)
    failed = sum(1 for a in assertions if not a.passed)
    critical_fails = [a.to_dict() for a in assertions if not a.passed and a.severity == "CRITICAL"]

    return {
        "status": "success",
        "sheet": sheet_name,
        "total_assertions": len(assertions),
        "passed": passed,
        "failed": failed,
        "pass_rate": f"{passed / len(assertions):.0%}",
        "critical_failures": critical_fails,
        "assertions": results,
    }


def main():
    parser = argparse.ArgumentParser(description="Validate DEODATE Lease Obligation Analysis workbook")
    parser.add_argument("workbook_path", help="Path to .xlsx file")
    parser.add_argument("--sheet", default=None, help="Sheet name (auto-detect if omitted)")
    args = parser.parse_args()

    result = validate(args.workbook_path, args.sheet)
    print(json.dumps(result, indent=2, default=str))

    if result.get("critical_failures"):
        print(f"\n⚠ {len(result['critical_failures'])} CRITICAL failure(s):")
        for cf in result["critical_failures"]:
            print(f"  [{cf['id']}] {cf['name']}: {cf['detail']}")


if __name__ == "__main__":
    main()
