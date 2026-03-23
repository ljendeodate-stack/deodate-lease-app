#!/usr/bin/env python3
"""
Rebuild T–W helper columns in the DEODATE Lease Obligation Analysis workbook.

Reads the active sheet's configuration (row bounds, input cell addresses, scenario
discount percentages), generates the correct formula strings for all schedule rows
in columns T–W, writes them via openpyxl, and optionally runs recalc.

Usage:
    python rebuild_helper_columns.py <workbook_path> [--sheet "Lease Schedule (2)"] [--dry-run] [--recalc]

Arguments:
    workbook_path   Path to the .xlsx file
    --sheet         Sheet name (default: auto-detect based on populated C21)
    --dry-run       Print formulas without writing
    --recalc        Run LibreOffice recalc after writing (requires recalc.py)
"""

import sys
import json
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ── Sheet configuration maps ──────────────────────────────────────────────────
# Each sheet variant has different cell addresses for the same logical inputs.

SHEET_CONFIGS = {
    "Lease Schedule": {
        "first_row": 38,
        "last_row": 161,
        "total_row": 162,
        "abatement_duration_cell": "$C$29",
        "abatement_proration_cell": "$C$32",
        "base_rent_col": "E",
        "month_num_col": "C",
        "nnn_total_col": "M",
        # Renegotiation discount % cells (row 6 in LS)
        "renego_discount_cells": {
            "T": "$K$6",   # Modest 10%
            "U": "$L$6",   # Material 20%
            "V": "$M$6",   # Significant 30%
        },
        # Exit buyout % cell
        "exit_discount_cell": "$N$24",  # Significant 50%
        "exit_col": "W",
    },
    "Lease Schedule (2)": {
        "first_row": 75,
        "last_row": 198,
        "total_row": 199,
        "abatement_duration_cell": "$C$37",
        "abatement_proration_cell": "$C$40",
        "base_rent_col": "E",
        "month_num_col": "C",
        "nnn_total_col": "M",
        # Renegotiation discount % cells (row 10 in LS2)
        "renego_discount_cells": {
            "T": "$G$10",  # Modest 10%
            "U": "$H$10",  # Material 20%
            "V": "$I$10",  # Significant 30%
        },
        # Exit buyout % cell
        "exit_discount_cell": "$J$29",  # Significant 50%
        "exit_col": "W",
    },
}


def detect_sheet(wb):
    """Auto-detect which sheet contains populated data by checking C21."""
    for name, cfg in SHEET_CONFIGS.items():
        if name in wb.sheetnames:
            ws = wb[name]
            c21 = ws["C21"].value
            if c21 is not None and c21 != 0:
                return name
    # Fallback: return the first configured sheet that exists
    for name in SHEET_CONFIGS:
        if name in wb.sheetnames:
            return name
    return None


def build_helper_formula(row, col_letter, cfg):
    """
    Build the formula string for a single helper column cell.

    Pattern:
        =IF(abatement_dur=0,
            base_rent*(1-discount%),
            IF(month#<=abatement_dur,
                0,
                IF(month#=abatement_dur+1,
                    base_rent*(1-discount%)* proration,
                    base_rent*(1-discount%)
                )
            )
        ) + NNN_total

    For renegotiation columns (T/U/V): discount applies to base rent only, NNN preserved.
    For exit column (W): same pattern but uses the exit buyout discount.
    """
    abat_dur = cfg["abatement_duration_cell"]
    abat_pro = cfg["abatement_proration_cell"]
    br = f'{cfg["base_rent_col"]}{row}'
    mn = f'{cfg["month_num_col"]}{row}'
    nnn = f'{cfg["nnn_total_col"]}{row}'

    if col_letter in cfg["renego_discount_cells"]:
        disc = cfg["renego_discount_cells"][col_letter]
    elif col_letter == cfg["exit_col"]:
        disc = cfg["exit_discount_cell"]
    else:
        raise ValueError(f"Unknown helper column: {col_letter}")

    formula = (
        f"=IF({abat_dur}=0,"
        f"{br}*(1-{disc}),"
        f"IF({mn}<={abat_dur},"
        f"0,"
        f"IF({mn}={abat_dur}+1,"
        f"{br}*(1-{disc})*{abat_pro},"
        f"{br}*(1-{disc})"
        f")))"
        f"+{nnn}"
    )
    return formula


def rebuild(workbook_path, sheet_name=None, dry_run=False, recalc=False):
    """Main rebuild logic."""
    wb_path = Path(workbook_path)
    if not wb_path.exists():
        print(f"Error: File not found: {wb_path}")
        return {"status": "error", "message": f"File not found: {wb_path}"}

    wb = load_workbook(str(wb_path))

    # Detect or validate sheet
    if sheet_name is None:
        sheet_name = detect_sheet(wb)
        if sheet_name is None:
            print("Error: Could not detect active sheet. Specify with --sheet.")
            return {"status": "error", "message": "Could not detect sheet"}
    elif sheet_name not in SHEET_CONFIGS:
        print(f"Error: Unknown sheet config for '{sheet_name}'.")
        print(f"Known sheets: {list(SHEET_CONFIGS.keys())}")
        return {"status": "error", "message": f"Unknown sheet: {sheet_name}"}

    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found in workbook.")
        return {"status": "error", "message": f"Sheet not found: {sheet_name}"}

    cfg = SHEET_CONFIGS[sheet_name]
    ws = wb[sheet_name]

    helper_cols = list(cfg["renego_discount_cells"].keys()) + [cfg["exit_col"]]
    first_row = cfg["first_row"]
    last_row = cfg["last_row"]
    total_rows = last_row - first_row + 1

    print(f"Sheet: {sheet_name}")
    print(f"Schedule rows: {first_row}–{last_row} ({total_rows} months)")
    print(f"Helper columns: {', '.join(helper_cols)}")
    print(f"Abatement duration cell: {cfg['abatement_duration_cell']}")
    print(f"Abatement proration cell: {cfg['abatement_proration_cell']}")
    print()

    # Check current state of T column
    t_first = ws[f"T{first_row}"].value
    t_second = ws[f"T{first_row + 1}"].value
    if t_first is not None and str(t_first).startswith("="):
        print(f"WARNING: T{first_row} already contains a formula: {t_first}")
        print("Helper columns may already be populated. Proceeding will overwrite.")
        print()

    formulas_written = 0
    sample_formulas = {}

    for col in helper_cols:
        for row in range(first_row, last_row + 1):
            formula = build_helper_formula(row, col, cfg)
            cell_ref = f"{col}{row}"

            if dry_run:
                if row == first_row:
                    sample_formulas[col] = formula
                    print(f"  {cell_ref}: {formula}")
                elif row == first_row + 1:
                    print(f"  {cell_ref}: {formula}")
                elif row == first_row + 2:
                    print(f"  ... ({total_rows - 2} more rows)")
            else:
                ws[cell_ref] = formula
                formulas_written += 1

        if not dry_run:
            sample_formulas[col] = build_helper_formula(first_row, col, cfg)

    if dry_run:
        print(f"\nDry run complete. Would write {total_rows * len(helper_cols)} formulas.")
        return {
            "status": "dry_run",
            "formulas_count": total_rows * len(helper_cols),
            "sample_formulas": sample_formulas,
        }

    # Save
    output_path = wb_path
    wb.save(str(output_path))
    print(f"\nWritten {formulas_written} formulas to {output_path}")

    # Optional recalc
    recalc_result = None
    if recalc:
        import subprocess
        recalc_script = Path(__file__).parent.parent.parent.parent / ".skills" / "skills" / "xlsx" / "scripts" / "recalc.py"
        if not recalc_script.exists():
            # Try alternate path
            recalc_script = Path("/sessions") / "inspiring-loving-ritchie" / "mnt" / ".skills" / "skills" / "xlsx" / "scripts" / "recalc.py"
        if recalc_script.exists():
            print(f"\nRunning recalc via {recalc_script}...")
            result = subprocess.run(
                ["python3", str(recalc_script), str(output_path)],
                capture_output=True, text=True, timeout=60
            )
            print(result.stdout)
            if result.stderr:
                print(f"Recalc stderr: {result.stderr}")
            try:
                recalc_result = json.loads(result.stdout.strip().split("\n")[-1])
            except (json.JSONDecodeError, IndexError):
                recalc_result = {"status": "unknown", "output": result.stdout}
        else:
            print(f"\nRecalc script not found at {recalc_script}. Skipping recalc.")

    return {
        "status": "success",
        "sheet": sheet_name,
        "formulas_written": formulas_written,
        "columns": helper_cols,
        "row_range": f"{first_row}–{last_row}",
        "sample_formulas": sample_formulas,
        "recalc_result": recalc_result,
    }


def main():
    parser = argparse.ArgumentParser(description="Rebuild T–W helper columns in lease schedule workbook")
    parser.add_argument("workbook_path", help="Path to .xlsx file")
    parser.add_argument("--sheet", default=None, help="Sheet name (auto-detect if omitted)")
    parser.add_argument("--dry-run", action="store_true", help="Print formulas without writing")
    parser.add_argument("--recalc", action="store_true", help="Run LibreOffice recalc after writing")
    args = parser.parse_args()

    result = rebuild(args.workbook_path, args.sheet, args.dry_run, args.recalc)
    print(f"\n{json.dumps(result, indent=2, default=str)}")


if __name__ == "__main__":
    main()
