#!/usr/bin/env python3
"""
Populate a blank DEODATE Lease Obligation Analysis workbook from a structured
input dictionary.

Writes all input values to the correct cells based on the target sheet
configuration, generates the monthly schedule date grid (Period Start / End,
Month #, Year #), and optionally triggers helper column rebuild and recalc.

Usage:
    python populate_template.py <workbook_path> <inputs_json_path> [--sheet "Lease Schedule (2)"] [--rebuild-tw] [--recalc]

The inputs JSON must conform to the schema defined in INPUTS_SCHEMA below.
Missing optional fields are left blank (not written). Missing required fields
cause an error with a clear message.

This script can also be imported and called programmatically:
    from populate_template import populate
    result = populate("workbook.xlsx", inputs_dict, sheet_name="Lease Schedule (2)")
"""

import sys
import json
import argparse
from pathlib import Path
from datetime import datetime, timedelta
import calendar

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl required. Install: pip install openpyxl")
    sys.exit(1)


# ── Input schema ──────────────────────────────────────────────────────────────

REQUIRED_FIELDS = [
    "rentable_sf",
    "lease_commencement_date",
    "total_lease_term_months",
    "base_rent_year1",
    "base_rent_escalation_rate",
]

OPTIONAL_FIELDS = [
    "rent_commencement_date",       # defaults to lease_commencement_date
    "lease_expiration_date",        # derived if omitted
    "effective_date_of_analysis",   # defaults to today
    "discount_rate",                # defaults to 0.07
    # NNN Year 1 Monthly Rates
    "cams_year1", "insurance_year1", "taxes_year1",
    "security_year1", "other_items_year1",
    # Escalation rates (default to base_rent_escalation_rate if omitted)
    "cams_escalation_rate", "insurance_escalation_rate",
    "taxes_escalation_rate", "security_escalation_rate",
    "other_items_escalation_rate",
    # Abatement
    "abatement_duration_months",    # defaults to 0
    "abatement_start_date",         # defaults to lease_commencement_date
    "abatement_end_date",           # derived from start + duration
    "abatement_pct",                # defaults to 1.0 (100%)
    "abatement_amount",             # derived from base_rent_year1
    "additional_abatement",         # "yes" or "no", defaults to "no"
    # NRC items (list of dicts)
    "nrc_items",                    # [{label, amount, date}]
    # Scenario parameters
    "renego_discounts",             # [0, 0.1, 0.2, 0.3]
    "exit_buyouts",                 # [0, 0.2, 0.3, 0.4, 0.5]
    "free_rent_months",             # defaults to 0
    "ti_psf",                       # defaults to 0
]


# ── Cell address maps ─────────────────────────────────────────────────────────

CELL_MAPS = {
    "Lease Schedule": {
        "title": "B2",
        "property_name": "B4",
        "rentable_sf": "C7",
        "lease_commencement_date": "C8",
        "rent_commencement_date": "C9",
        "total_lease_term_months": "C10",
        "lease_expiration_date": "C11",
        "effective_date_of_analysis": "C12",
        "base_rent_year1": "C21",
        "insurance_year1": "C22",
        "taxes_year1": "C23",
        "cams_year1": "C24",
        "security_year1": "C25",
        "other_items_year1": "C26",
        "base_rent_escalation_rate": "F5",
        "cams_escalation_rate": "F6",
        "insurance_escalation_rate": "F7",
        "taxes_escalation_rate": "F8",
        "security_escalation_rate": "F9",
        "other_items_escalation_rate": "F10",
        "abatement_duration_months": "C29",
        "abatement_start_date": "C30",
        "abatement_end_date": "C31",
        "abatement_pct": "C32",
        "abatement_amount": "C33",
        "additional_abatement": "C34",
        "discount_rate": "J2",
        "free_rent_months": "J14",
        "ti_psf": "J15",
        # NRC table
        "nrc_start_row": 14,
        "nrc_amount_col": "F",
        "nrc_date_col": "G",
        "nrc_label_col": "E",
        # Scenario discount cells
        "renego_discount_cells": ["J6", "K6", "L6", "M6"],
        "exit_buyout_cells": ["J24", "K24", "L24", "M24", "N24"],
        # Schedule
        "schedule_first_row": 38,
        "schedule_last_row": 161,
    },
    "Lease Schedule (2)": {
        "title": "B2",
        "property_name": "B4",
        "rentable_sf": "C7",
        "lease_commencement_date": "C8",
        "rent_commencement_date": "C9",
        "total_lease_term_months": "C10",
        "lease_expiration_date": "C11",
        "effective_date_of_analysis": "C12",
        "base_rent_year1": "C21",
        "insurance_year1": "C22",
        "taxes_year1": "C23",
        "cams_year1": "C24",
        "security_year1": "C25",
        "other_items_year1": "C26",
        "base_rent_escalation_rate": "C29",
        "cams_escalation_rate": "C30",
        "insurance_escalation_rate": "C31",
        "taxes_escalation_rate": "C32",
        "security_escalation_rate": "C33",
        "other_items_escalation_rate": "C34",
        "abatement_duration_months": "C37",
        "abatement_start_date": "C38",
        "abatement_end_date": "C39",
        "abatement_pct": "C40",
        "abatement_amount": "C41",
        "additional_abatement": "C42",
        "discount_rate": "F7",
        "free_rent_months": "F18",
        "ti_psf": "F19",
        # NRC table
        "nrc_start_row": 53,
        "nrc_amount_col": "C",
        "nrc_date_col": "D",
        "nrc_label_col": "B",
        # Scenario discount cells
        "renego_discount_cells": ["F10", "G10", "H10", "I10"],
        "exit_buyout_cells": ["F29", "G29", "H29", "I29", "J29"],
        # Schedule
        "schedule_first_row": 75,
        "schedule_last_row": 198,
    },
}


def parse_date(val):
    """Parse a date from string, datetime, or leave as-is."""
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
        raise ValueError(f"Cannot parse date: {val}")
    return val


def add_months(dt, months):
    """Add N months to a date, landing on the 1st of the target month."""
    month = dt.month - 1 + months
    year = dt.year + month // 12
    month = month % 12 + 1
    return datetime(year, month, 1)


def end_of_month(dt):
    """Return last day of the month for a given date."""
    last_day = calendar.monthrange(dt.year, dt.month)[1]
    return datetime(dt.year, dt.month, last_day)


def generate_schedule_dates(commencement, term_months):
    """Generate period start/end dates, month#, year# for the full schedule."""
    rows = []
    for i in range(term_months):
        period_start = add_months(commencement, i)
        period_end = end_of_month(period_start)
        month_num = i + 1
        year_num = (i // 12) + 1
        rows.append({
            "period_start": period_start,
            "period_end": period_end,
            "month_num": month_num,
            "year_num": year_num,
        })
    return rows


def apply_defaults(inputs):
    """Fill in optional fields with sensible defaults."""
    inp = dict(inputs)

    if "rent_commencement_date" not in inp or inp["rent_commencement_date"] is None:
        inp["rent_commencement_date"] = inp["lease_commencement_date"]

    commencement = parse_date(inp["lease_commencement_date"])
    term = int(inp["total_lease_term_months"])

    if "lease_expiration_date" not in inp or inp["lease_expiration_date"] is None:
        exp = add_months(commencement, term - 1)
        inp["lease_expiration_date"] = end_of_month(exp)

    if "effective_date_of_analysis" not in inp or inp["effective_date_of_analysis"] is None:
        inp["effective_date_of_analysis"] = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    if "discount_rate" not in inp or inp["discount_rate"] is None:
        inp["discount_rate"] = 0.07

    # NNN defaults
    for field in ["cams_year1", "insurance_year1", "taxes_year1", "security_year1", "other_items_year1"]:
        if field not in inp or inp[field] is None:
            inp[field] = 0

    # Escalation defaults
    base_esc = inp["base_rent_escalation_rate"]
    for field in ["cams_escalation_rate", "insurance_escalation_rate", "taxes_escalation_rate"]:
        if field not in inp or inp[field] is None:
            inp[field] = base_esc
    for field in ["security_escalation_rate", "other_items_escalation_rate"]:
        if field not in inp or inp[field] is None:
            inp[field] = 0.02  # Default 2% for security and other

    # Abatement defaults
    if "abatement_duration_months" not in inp or inp["abatement_duration_months"] is None:
        inp["abatement_duration_months"] = 0
    abat_dur = int(inp["abatement_duration_months"])

    if abat_dur > 0:
        if "abatement_start_date" not in inp or inp["abatement_start_date"] is None:
            inp["abatement_start_date"] = commencement
        abat_start = parse_date(inp["abatement_start_date"])
        if "abatement_end_date" not in inp or inp["abatement_end_date"] is None:
            end_month = add_months(abat_start, abat_dur - 1)
            inp["abatement_end_date"] = end_of_month(end_month)
        if "abatement_pct" not in inp or inp["abatement_pct"] is None:
            inp["abatement_pct"] = 1.0
        if "abatement_amount" not in inp or inp["abatement_amount"] is None:
            inp["abatement_amount"] = inp["base_rent_year1"]
    else:
        inp.setdefault("abatement_start_date", None)
        inp.setdefault("abatement_end_date", None)
        inp.setdefault("abatement_pct", 0)
        inp.setdefault("abatement_amount", 0)

    inp.setdefault("additional_abatement", "no")
    inp.setdefault("free_rent_months", 0)
    inp.setdefault("ti_psf", 0)
    inp.setdefault("renego_discounts", [0, 0.1, 0.2, 0.3])
    inp.setdefault("exit_buyouts", [0, 0.2, 0.3, 0.4, 0.5])
    inp.setdefault("nrc_items", [])

    return inp


def validate_inputs(inputs):
    """Check required fields and return list of errors."""
    errors = []
    for field in REQUIRED_FIELDS:
        if field not in inputs or inputs[field] is None:
            errors.append(f"Missing required field: {field}")

    if not errors:
        try:
            commencement = parse_date(inputs["lease_commencement_date"])
            term = int(inputs["total_lease_term_months"])
            if term < 1 or term > 600:
                errors.append(f"Lease term {term} months is out of plausible range (1–600)")
            sf = float(inputs["rentable_sf"])
            if sf <= 0:
                errors.append(f"Rentable SF must be positive, got {sf}")
            rent = float(inputs["base_rent_year1"])
            if rent < 0:
                errors.append(f"Base rent cannot be negative: {rent}")
            esc = float(inputs["base_rent_escalation_rate"])
            if esc < 0 or esc > 0.20:
                errors.append(f"Escalation rate {esc:.1%} is outside plausible range (0–20%)")
        except (ValueError, TypeError) as e:
            errors.append(f"Invalid input value: {e}")

    return errors


def populate(workbook_path, inputs, sheet_name=None, rebuild_tw=False, recalc=False):
    """
    Populate the workbook with input values and generate schedule dates.

    Returns a result dict with status, cells_written, warnings, and errors.
    """
    wb_path = Path(workbook_path)
    if not wb_path.exists():
        return {"status": "error", "message": f"File not found: {wb_path}"}

    # Validate
    errors = validate_inputs(inputs)
    if errors:
        return {"status": "error", "validation_errors": errors}

    # Apply defaults
    inp = apply_defaults(inputs)

    # Load workbook
    wb = load_workbook(str(wb_path))

    # Determine sheet
    if sheet_name is None:
        sheet_name = "Lease Schedule (2)" if "Lease Schedule (2)" in wb.sheetnames else "Lease Schedule"
    if sheet_name not in wb.sheetnames:
        return {"status": "error", "message": f"Sheet '{sheet_name}' not found"}
    if sheet_name not in CELL_MAPS:
        return {"status": "error", "message": f"No cell map for sheet '{sheet_name}'"}

    cmap = CELL_MAPS[sheet_name]
    ws = wb[sheet_name]
    cells_written = 0
    warnings = []

    # ── Write scalar inputs ──────────────────────────────────────────────
    scalar_fields = [
        "rentable_sf", "lease_commencement_date", "rent_commencement_date",
        "total_lease_term_months", "lease_expiration_date", "effective_date_of_analysis",
        "base_rent_year1", "insurance_year1", "taxes_year1", "cams_year1",
        "security_year1", "other_items_year1",
        "base_rent_escalation_rate", "cams_escalation_rate", "insurance_escalation_rate",
        "taxes_escalation_rate", "security_escalation_rate", "other_items_escalation_rate",
        "abatement_duration_months", "abatement_pct", "abatement_amount",
        "additional_abatement", "discount_rate", "free_rent_months", "ti_psf",
    ]

    for field in scalar_fields:
        if field in cmap and field in inp and inp[field] is not None:
            cell_ref = cmap[field]
            val = inp[field]
            # Parse dates
            if "date" in field:
                val = parse_date(val)
            ws[cell_ref] = val
            cells_written += 1

    # Abatement dates (separate handling for 0-month abatement)
    if inp["abatement_duration_months"] > 0:
        for date_field in ["abatement_start_date", "abatement_end_date"]:
            if date_field in cmap and inp.get(date_field):
                ws[cmap[date_field]] = parse_date(inp[date_field])
                cells_written += 1

    # ── Write scenario discount percentages ──────────────────────────────
    renego = inp.get("renego_discounts", [0, 0.1, 0.2, 0.3])
    for i, cell_ref in enumerate(cmap["renego_discount_cells"]):
        if i < len(renego):
            ws[cell_ref] = renego[i]
            cells_written += 1

    exits = inp.get("exit_buyouts", [0, 0.2, 0.3, 0.4, 0.5])
    for i, cell_ref in enumerate(cmap["exit_buyout_cells"]):
        if i < len(exits):
            ws[cell_ref] = exits[i]
            cells_written += 1

    # ── Write NRC items ──────────────────────────────────────────────────
    nrc_items = inp.get("nrc_items", [])
    nrc_start = cmap["nrc_start_row"]
    for i, item in enumerate(nrc_items):
        row = nrc_start + i
        if "label" in item and item["label"]:
            ws[f'{cmap["nrc_label_col"]}{row}'] = item["label"]
        if "amount" in item and item["amount"] is not None:
            ws[f'{cmap["nrc_amount_col"]}{row}'] = item["amount"]
            cells_written += 1
        if "date" in item and item["date"]:
            try:
                ws[f'{cmap["nrc_date_col"]}{row}'] = parse_date(item["date"])
                cells_written += 1
            except (ValueError, TypeError):
                ws[f'{cmap["nrc_date_col"]}{row}'] = str(item["date"])
                warnings.append(f"NRC row {row}: date '{item['date']}' stored as text — will be excluded from OTC Remaining")
                cells_written += 1

    # ── Generate schedule dates ──────────────────────────────────────────
    commencement = parse_date(inp["lease_commencement_date"])
    term = int(inp["total_lease_term_months"])
    schedule_rows = generate_schedule_dates(commencement, term)
    first_row = cmap["schedule_first_row"]

    max_row = cmap["schedule_last_row"]
    if len(schedule_rows) > (max_row - first_row + 1):
        warnings.append(
            f"Term ({term} months) exceeds template capacity ({max_row - first_row + 1} rows). "
            f"Schedule truncated at row {max_row}."
        )

    for i, srow in enumerate(schedule_rows):
        row = first_row + i
        if row > max_row:
            break
        ws[f"A{row}"] = srow["period_start"]
        ws[f"B{row}"] = srow["period_end"]
        ws[f"C{row}"] = srow["month_num"]
        ws[f"D{row}"] = srow["year_num"]
        cells_written += 4

    # ── Save ─────────────────────────────────────────────────────────────
    wb.save(str(wb_path))
    print(f"Populated {cells_written} cells in '{sheet_name}'")

    # ── Optional: rebuild T-W helper columns ─────────────────────────────
    rebuild_result = None
    if rebuild_tw:
        try:
            from rebuild_helper_columns import rebuild
            rebuild_result = rebuild(str(wb_path), sheet_name, dry_run=False, recalc=False)
        except ImportError:
            script_dir = Path(__file__).parent
            sys.path.insert(0, str(script_dir))
            try:
                from rebuild_helper_columns import rebuild
                rebuild_result = rebuild(str(wb_path), sheet_name, dry_run=False, recalc=False)
            except ImportError:
                warnings.append("Could not import rebuild_helper_columns.py for T-W rebuild")

    # ── Optional: recalc ─────────────────────────────────────────────────
    recalc_result = None
    if recalc:
        import subprocess
        recalc_script = Path(__file__).parent.parent.parent.parent / ".skills" / "skills" / "xlsx" / "scripts" / "recalc.py"
        if not recalc_script.exists():
            recalc_script = Path("/sessions") / "inspiring-loving-ritchie" / "mnt" / ".skills" / "skills" / "xlsx" / "scripts" / "recalc.py"
        if recalc_script.exists():
            result = subprocess.run(
                ["python3", str(recalc_script), str(wb_path)],
                capture_output=True, text=True, timeout=60
            )
            try:
                recalc_result = json.loads(result.stdout.strip().split("\n")[-1])
            except (json.JSONDecodeError, IndexError):
                recalc_result = {"status": "unknown", "output": result.stdout}

    return {
        "status": "success",
        "sheet": sheet_name,
        "cells_written": cells_written,
        "schedule_rows_generated": min(len(schedule_rows), max_row - first_row + 1),
        "warnings": warnings,
        "rebuild_result": rebuild_result,
        "recalc_result": recalc_result,
    }


def main():
    parser = argparse.ArgumentParser(description="Populate lease schedule template from inputs JSON")
    parser.add_argument("workbook_path", help="Path to .xlsx file")
    parser.add_argument("inputs_json", help="Path to JSON file with input values")
    parser.add_argument("--sheet", default=None, help="Sheet name (auto-detect if omitted)")
    parser.add_argument("--rebuild-tw", action="store_true", help="Rebuild T-W helper columns after populating")
    parser.add_argument("--recalc", action="store_true", help="Run LibreOffice recalc after writing")
    args = parser.parse_args()

    with open(args.inputs_json) as f:
        inputs = json.load(f)

    result = populate(args.workbook_path, inputs, args.sheet, args.rebuild_tw, args.recalc)
    print(f"\n{json.dumps(result, indent=2, default=str)}")


if __name__ == "__main__":
    main()
